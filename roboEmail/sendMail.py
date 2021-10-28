
""" Nome: sendMail """
""" Autor: Felipe Lamim <felipe.lamim@bradescoseguros.com.br> """
""" Descrição: Envia e-mail automáticamente para os grupos de projeto da BARE, informando os Incidentes pendentes em seus grupos """
import selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import numpy as np
import pandas as pd
from pandas import ExcelWriter
from pandas import ExcelFile
import csv
from xlsxwriter.workbook import Workbook
from openpyxl import Workbook as WB
from openpyxl import load_workbook
from statistics import mode
import os
import glob

gruposFile = Workbook()
caminhoGrupos = 'C:\\Users\\g571676\\Desktop\\responsaveis_grupos.xlsx'
gruposFile = load_workbook(caminhoGrupos)
planilhaGroups = gruposFile.active
max_linha = planilhaGroups.max_row
max_coluna = planilhaGroups.max_column
df = pd.read_excel(caminhoGrupos, sheetname='Plan1')
grupos = df['Grupo'].values
responsavel = df['Responsavel'].values
print("Iniciando execução...")

def instancia_driver():    #INICIALIZACAO
    options = webdriver.ChromeOptions()
    options.add_argument('--ignore-certificate-errors')
    options.add_argument("--test-type")
    options.add_argument("--start-maximized")
    prefs = {
        "download.default_directory": "\\\\srv-arquivos07\dirgerti\SEDT_AUTORE\TN_AUTORE\SUSTENTAÇÃO\Incidentes\Dashboard Incidentes",
        "download.prompt_for_download": False,
        "download.directory_upgrade": True
    }
    options.add_experimental_option('prefs', prefs)

    driver = webdriver.Chrome('C:\\Users\\g571676\\Documents\\Python_Scripts\\roboEmail\\chromedriver.exe', options=options)
    return driver

def clica_id(driver, id):
    element = WebDriverWait(driver, 20).until(
    EC.element_to_be_clickable((By.ID, id)))
    element.click()

def clica_xpath(driver, xpath):
    element = WebDriverWait(driver, 20).until(
    EC.element_to_be_clickable((By.XPATH, xpath)))
    element.click()

#preenche texto
def insere_texto(elemento, texto):
    elemento.send_keys(texto)
#clica no botão "Novo"
def newMsg(driver):
    clica_id(driver, 'newmsgc') #abre a janela para um novo email

def preencheXpath(driver, xpath, texto):
    element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, xpath))
    )
    element.send_keys(texto)


def retorna_objetos(driver, selecao, nome):
    if (selecao == 'id'):
        return driver.find_element_by_id(nome)
    if (selecao == 'class'):
        return driver.find_elements_by_class_name(nome)
    if (selecao == 'name'):
        return driver.find_element_by_name(nome)
    if ((selecao == 'tag')):
        return driver.find_elements_by_tag_name(nome)

#filtra pelo grupo solucionador
def filtra(driver, grupo):
    arquivo_excel = Workbook()
    caminho = '\\\\srv-arquivos07\\DirGerTI\\SEDT_AUTORE\\TN_AUTORE\\SUSTENTAÇÃO\\Incidentes\\Dashboard Incidentes\\base dashboard incidentes.xlsx'
    arquivo_excel = load_workbook(caminho)
    planilha = arquivo_excel.active
    df = pd.read_excel(caminho, sheetname='Planilha1')
    df = df[df['Designação principal'] == grupo]
    status = ['DIRECIONADO', 'EM TRATAMENTO']
    df = df[df['Status'].isin(status)]
    iNs = df['ID do Incidente'].values
    dtAbertura = df['Hora de Abertura'].values
    desc = df['Descrição Resumida'].values
    print('-- Filtrando por grupos --')
    print('=> ' + grupo)
    print('Total de ' + str(len(iNs)) + ' Incidentes encontrados')
     ##########INICIO DA CRIAÇÃO DA PLANILHA########
    print('Criando arquivo...')
    fileName = 'IncidentesPendentes' + str(i) + '.xlsx'
    caminhoIn = 'C:\\Users\\g571676\\Documents\\Python_Scripts\\roboEmail\\' + fileName
    df.to_excel(caminhoIn, 'Planilha1', index=False)
     ###########FIM DA CRIAÇÃO DA PLANILHA#########
           
    #####Anexar a planilha######
    print('anexando arquivo...')
    clica_id(driver, 'attachfile')
    driver.switch_to_frame('iFrameModalDlg') #entra no primeiro frame
    time.sleep(5)
    driver.switch_to.frame(driver.find_elements_by_tag_name("iframe")[0]) #entra no segundo 
    fileName = 'IncidentesPendentes' + str(i)
    driver.find_element_by_id("file1").send_keys(os.getcwd()+"\\"+fileName+'.xlsx')
    clica_id(driver, 'btnAttch')
    print('Arquivo anexado...')
    #####Fim Anexar a planilha######

#acessa o arquivo onde estão os grupos e seus respectivos responsáveis
def leGrupos(driver):
    gruposFile = Workbook()
    caminhoGrupos = 'C:\\Users\\g571676\\Desktop\\responsaveis_grupos.xlsx'
    gruposFile = load_workbook(caminhoGrupos)
    planilhaGroups = gruposFile.active
    max_linha = planilhaGroups.max_row
    max_coluna = planilhaGroups.max_column
    df = pd.read_excel(caminhoGrupos, sheetname='Plan1')
    grupos = df['Grupo'].values
    responsavel = df['Responsavel'].values
    newMsg(driver)       
    preenche(driver, responsavel[i])
    filtra(driver, grupos[i])
#preenche o e-mail
def preenche(driver, destinatario):
    window_before = driver.window_handles[0]
    window_after = driver.window_handles[1]
    driver.switch_to_window(window_after)#muda de janela
    insere_texto(retorna_objetos(driver, 'id', 'divTo'), destinatario) #preenche campo 'Para'
    insere_texto(retorna_objetos(driver, 'id', 'txtSubj'), '[EMAIL AUTOMÁTICO - SUSTENTAÇÃO BARE] Incidentes Pendentes') #preenche campo 'assunto'

    #texto base do e-mail   
    textoBase = 'Prezado(a), seguem em anexo os INs pendentes para o seu grupo'
    
    insere_texto(retorna_objetos(driver, 'id', 'ifBdy'), textoBase + ': \n\n')
    clica_id(driver, 'imphigh')
    
#clica em enviar        
def send(driver):
    
    clica_id(driver, 'send')
    print('Enviando..')
    time.sleep(5)
    driver.quit()

def main_run():
    driver = instancia_driver()
    driver.get('https://webmail10.bradseg.com.br/owa/')
    leGrupos(driver)
    send(driver)

    return driver
  

i = 0
while(i != len(grupos)):
    driver = main_run()
    i += 1
print('-- Fim da execução --')